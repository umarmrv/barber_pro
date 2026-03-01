from __future__ import annotations

from datetime import UTC, date, datetime, timedelta
from io import BytesIO

import pytest

from barber_bot.db.repositories import TodayBookingDetailed
from barber_bot.services.export_xlsx import (
    build_booking_stats_filename,
    build_booking_stats_workbook_bytes,
)

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:  # pragma: no cover - depends on local env deps
    load_workbook = None


def test_build_booking_stats_workbook_bytes_contains_summary_and_manual_columns() -> None:
    if load_workbook is None:
        pytest.skip("openpyxl is not installed")

    starts = datetime(2026, 2, 22, 11, 30, tzinfo=UTC)
    booking = TodayBookingDetailed(
        booking_id=17,
        starts_at_utc=starts,
        ends_at_utc=starts + timedelta(minutes=30),
        status="completed",
        barber_id=1,
        barber_name="Ali",
        service_id=1,
        service_name_ru="Стрижка",
        service_name_uz="Soch",
        service_name_tj="Мӯйсаргирӣ",
        client_id=10,
        client_tg_user_id=123456789,
        client_tg_username="client_1",
        client_phone_e164="+998901234567",
    )

    payload = build_booking_stats_workbook_bytes(
        locale="ru",
        tz_name="UTC",
        period_from_local=date(2026, 2, 22),
        period_to_local=date(2026, 2, 28),
        total=1,
        confirmed=0,
        completed=1,
        cancelled=0,
        blocked=0,
        cash_total_minor=12345,
        bookings=[booking],
    )

    wb = load_workbook(BytesIO(payload))
    ws = wb["Stats"]

    assert ws["A1"].value == "Период"
    assert ws["B2"].value == 1
    assert ws["A4"].value == "Обслужено"
    assert ws["B4"].value == 1
    assert ws["A7"].value == "Касса (обслужено), TJS"
    assert float(ws["B7"].value) == 123.45
    assert ws.cell(row=9, column=13).value == "Факт визита (ручной ввод)"
    assert ws.cell(row=9, column=14).value == "Комментарий админа"
    assert ws.cell(row=10, column=4).value == "обслужена"
    assert ws.cell(row=10, column=13).value is None
    assert ws.cell(row=10, column=14).value is None


def test_build_booking_stats_filename() -> None:
    now_local = datetime(2026, 2, 22, 12, 30)
    week_name = build_booking_stats_filename(mode="week", now_local=now_local)
    date_name = build_booking_stats_filename(
        mode="date",
        now_local=now_local,
        local_day=(now_local + timedelta(days=1)).date(),
    )

    assert week_name.startswith("booking_stats_week_20260222_123000")
    assert week_name.endswith(".xlsx")
    assert date_name == "booking_stats_date_20260223.xlsx"


def test_build_booking_stats_workbook_bytes_tj_uses_tajik_service_name() -> None:
    if load_workbook is None:
        pytest.skip("openpyxl is not installed")

    starts = datetime(2026, 2, 22, 11, 30, tzinfo=UTC)
    booking = TodayBookingDetailed(
        booking_id=1,
        starts_at_utc=starts,
        ends_at_utc=starts + timedelta(minutes=30),
        status="confirmed",
        barber_id=1,
        barber_name="Usto",
        service_id=1,
        service_name_ru="Стрижка",
        service_name_uz="Soch",
        service_name_tj="Мӯйсаргирӣ",
        client_id=1,
        client_tg_user_id=42,
        client_tg_username="tj_user",
        client_phone_e164="+992900000001",
    )

    payload = build_booking_stats_workbook_bytes(
        locale="tj",
        tz_name="UTC",
        period_from_local=date(2026, 2, 22),
        period_to_local=date(2026, 2, 22),
        total=1,
        confirmed=1,
        completed=0,
        cancelled=0,
        blocked=0,
        cash_total_minor=0,
        bookings=[booking],
    )

    wb = load_workbook(BytesIO(payload))
    ws = wb["Stats"]
    assert ws.cell(row=9, column=6).value == "Хизмат"
    assert ws.cell(row=10, column=6).value == "Мӯйсаргирӣ"
