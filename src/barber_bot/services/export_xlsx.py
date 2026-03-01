from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from zoneinfo import ZoneInfo

from barber_bot.db.repositories import TodayBookingDetailed


_LABELS: dict[str, dict[str, str]] = {
    "ru": {
        "sheet_title": "Stats",
        "period": "Период",
        "total": "Всего записей",
        "confirmed": "Подтверждено",
        "completed": "Обслужено",
        "cancelled": "Отменено",
        "blocked": "Блокировок",
        "cash_total": "Касса (обслужено), TJS",
        "h_booking_id": "Booking ID",
        "h_local_date": "Локальная дата",
        "h_local_time": "Локальное время",
        "h_status": "Статус",
        "h_barber": "Мастер",
        "h_service": "Услуга",
        "h_username": "Клиент (username)",
        "h_tg_user_id": "TG user id",
        "h_phone": "Телефон",
        "h_client_id": "Client ID",
        "h_starts_utc": "Starts UTC",
        "h_ends_utc": "Ends UTC",
        "h_manual_visit": "Факт визита (ручной ввод)",
        "h_manual_note": "Комментарий админа",
        "status_confirmed": "подтверждена",
        "status_completed": "обслужена",
        "status_cancelled": "отменена",
        "status_blocked": "блок",
    },
    "uz": {
        "sheet_title": "Stats",
        "period": "Davr",
        "total": "Jami yozuvlar",
        "confirmed": "Tasdiqlangan",
        "completed": "Xizmat ko'rsatilgan",
        "cancelled": "Bekor qilingan",
        "blocked": "Bandlangan",
        "cash_total": "Kassa (xizmat), TJS",
        "h_booking_id": "Booking ID",
        "h_local_date": "Mahalliy sana",
        "h_local_time": "Mahalliy vaqt",
        "h_status": "Holat",
        "h_barber": "Usta",
        "h_service": "Xizmat",
        "h_username": "Mijoz (username)",
        "h_tg_user_id": "TG user id",
        "h_phone": "Telefon",
        "h_client_id": "Client ID",
        "h_starts_utc": "Starts UTC",
        "h_ends_utc": "Ends UTC",
        "h_manual_visit": "Tashrif holati (qo'lda)",
        "h_manual_note": "Admin izohi",
        "status_confirmed": "tasdiqlangan",
        "status_completed": "xizmat ko'rsatilgan",
        "status_cancelled": "bekor qilingan",
        "status_blocked": "bandlangan",
    },
    "tj": {
        "sheet_title": "Stats",
        "period": "Давра",
        "total": "Ҳамагӣ сабтҳо",
        "confirmed": "Тасдиқшуда",
        "completed": "Хизмат расонида шуд",
        "cancelled": "Бекоршуда",
        "blocked": "Бандшуда",
        "cash_total": "Касса (хизмат), TJS",
        "h_booking_id": "Booking ID",
        "h_local_date": "Санаи маҳаллӣ",
        "h_local_time": "Вақти маҳаллӣ",
        "h_status": "Ҳолат",
        "h_barber": "Усто",
        "h_service": "Хизмат",
        "h_username": "Мизоҷ (username)",
        "h_tg_user_id": "TG user id",
        "h_phone": "Телефон",
        "h_client_id": "Client ID",
        "h_starts_utc": "Starts UTC",
        "h_ends_utc": "Ends UTC",
        "h_manual_visit": "Ҳолати ташриф (дастӣ)",
        "h_manual_note": "Шарҳи админ",
        "status_confirmed": "тасдиқшуда",
        "status_completed": "хизмат расонида шуд",
        "status_cancelled": "бекоршуда",
        "status_blocked": "бандшуда",
    },
}


def _l(locale: str, key: str) -> str:
    lang = locale if locale in _LABELS else "ru"
    return _LABELS[lang].get(key, _LABELS["ru"].get(key, key))


def _status_label(locale: str, status: str) -> str:
    key = f"status_{status}"
    value = _l(locale, key)
    return status if value == key else value


def build_booking_stats_workbook_bytes(
    *,
    locale: str,
    tz_name: str,
    period_from_local: date,
    period_to_local: date,
    total: int,
    confirmed: int,
    completed: int,
    cancelled: int,
    blocked: int,
    cash_total_minor: int,
    bookings: list[TodayBookingDetailed],
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = _l(locale, "sheet_title")

    ws["A1"] = _l(locale, "period")
    ws["B1"] = f"{period_from_local.isoformat()} - {period_to_local.isoformat()}"
    ws["A2"] = _l(locale, "total")
    ws["B2"] = total
    ws["A3"] = _l(locale, "confirmed")
    ws["B3"] = confirmed
    ws["A4"] = _l(locale, "completed")
    ws["B4"] = completed
    ws["A5"] = _l(locale, "cancelled")
    ws["B5"] = cancelled
    ws["A6"] = _l(locale, "blocked")
    ws["B6"] = blocked
    ws["A7"] = _l(locale, "cash_total")
    ws["B7"] = cash_total_minor / 100

    headers = [
        _l(locale, "h_booking_id"),
        _l(locale, "h_local_date"),
        _l(locale, "h_local_time"),
        _l(locale, "h_status"),
        _l(locale, "h_barber"),
        _l(locale, "h_service"),
        _l(locale, "h_username"),
        _l(locale, "h_tg_user_id"),
        _l(locale, "h_phone"),
        _l(locale, "h_client_id"),
        _l(locale, "h_starts_utc"),
        _l(locale, "h_ends_utc"),
        _l(locale, "h_manual_visit"),
        _l(locale, "h_manual_note"),
    ]

    header_row = 9
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True)

    tz = ZoneInfo(tz_name)
    for offset, booking in enumerate(bookings, start=1):
        row = header_row + offset
        local_start = booking.starts_at_utc.astimezone(tz)
        local_date = local_start.strftime("%Y-%m-%d")
        local_time = local_start.strftime("%H:%M")
        if locale == "uz":
            service_name = booking.service_name_uz or booking.service_name_ru or booking.service_name_tj or "-"
        elif locale == "tj":
            service_name = booking.service_name_tj or booking.service_name_ru or booking.service_name_uz or "-"
        else:
            service_name = booking.service_name_ru or booking.service_name_uz or booking.service_name_tj or "-"

        username = "-"
        if booking.client_tg_username:
            username = f"@{booking.client_tg_username}"
        elif booking.client_tg_user_id:
            username = str(booking.client_tg_user_id)

        ws.cell(row=row, column=1, value=booking.booking_id)
        ws.cell(row=row, column=2, value=local_date)
        ws.cell(row=row, column=3, value=local_time)
        ws.cell(row=row, column=4, value=_status_label(locale, booking.status))
        ws.cell(row=row, column=5, value=booking.barber_name or "-")
        ws.cell(row=row, column=6, value=service_name)
        ws.cell(row=row, column=7, value=username)
        ws.cell(row=row, column=8, value=booking.client_tg_user_id or "")
        ws.cell(row=row, column=9, value=booking.client_phone_e164 or "-")
        ws.cell(row=row, column=10, value=booking.client_id or "")
        ws.cell(
            row=row,
            column=11,
            value=booking.starts_at_utc.isoformat().replace("+00:00", "Z"),
        )
        ws.cell(
            row=row,
            column=12,
            value=booking.ends_at_utc.isoformat().replace("+00:00", "Z"),
        )
        ws.cell(row=row, column=13, value="")
        ws.cell(row=row, column=14, value="")

    for column in ws.columns:
        max_len = 0
        column_letter = column[0].column_letter
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        ws.column_dimensions[column_letter].width = min(max_len + 2, 48)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_booking_stats_filename(
    *,
    mode: str,
    now_local: datetime,
    local_day: date | None = None,
) -> str:
    if mode == "date" and local_day is not None:
        return f"booking_stats_date_{local_day.strftime('%Y%m%d')}.xlsx"
    stamp = now_local.strftime("%Y%m%d_%H%M%S")
    return f"booking_stats_{mode}_{stamp}.xlsx"
